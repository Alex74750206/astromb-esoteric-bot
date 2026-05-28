"""Экспорт базы клиентов в Excel для рассылок и аналитики."""
import os
import sqlite3
import asyncio
import logging
import shutil
import tempfile
from datetime import datetime

log = logging.getLogger(__name__)

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import DB_PATH

# Файл сохраняется рядом с БД (локально — в папке эзотерика/, на fly.io — на volume)
_BOT_DIR   = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))   # bot/
_ROOT_DIR  = os.path.dirname(_BOT_DIR)                                      # эзотерика/
EXCEL_PATH = os.path.join(os.path.dirname(os.path.abspath(DB_PATH)) or _ROOT_DIR, "Клиенты.xlsx")

_HEADER_COLOR = "1A3C5E"
_ROW_ALT      = "EBF2FA"
_GREEN        = "1E8449"
_ORANGE       = "CA6F1E"


def _thin():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)


def _sync_excel():
    """Синхронная запись — вызывается через asyncio.to_thread."""
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row

    users = conn.execute(
        "SELECT user_id, username, full_name, birth_date, created_at, funnel_step "
        "FROM users ORDER BY created_at DESC"
    ).fetchall()

    purch_rows = conn.execute(
        "SELECT user_id, COUNT(*) as cnt, COALESCE(SUM(stars_paid),0) as stars "
        "FROM purchases GROUP BY user_id"
    ).fetchall()
    conn.close()

    purch = {r["user_id"]: (r["cnt"], r["stars"]) for r in purch_rows}

    wb = openpyxl.Workbook()

    # ── Лист «Клиенты» ───────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Клиенты"
    ws.freeze_panes = "A2"

    headers = [
        "ID Telegram", "Username", "Имя в Telegram",
        "Дата рождения", "Дата регистрации", "Шаг воронки",
        "Покупок", "Stars потрачено",
    ]
    col_widths = [16, 22, 26, 16, 22, 14, 10, 16]

    hdr_font  = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    hdr_fill  = PatternFill(start_color=_HEADER_COLOR, end_color=_HEADER_COLOR, fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center")

    for col_i, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_i, value=h)
        cell.font  = hdr_font
        cell.fill  = hdr_fill
        cell.alignment = hdr_align
        cell.border = _thin()
        ws.column_dimensions[get_column_letter(col_i)].width = w

    ws.row_dimensions[1].height = 22

    for row_i, user in enumerate(users, 2):
        uid = user["user_id"]
        cnt, stars = purch.get(uid, (0, 0))

        row_data = [
            uid,
            f"@{user['username']}" if user["username"] else "—",
            user["full_name"] or "—",
            user["birth_date"] or "не указана",
            (user["created_at"] or "")[:16].replace("T", " "),
            user["funnel_step"] or 0,
            cnt,
            stars,
        ]

        fill = PatternFill(start_color=_ROW_ALT, end_color=_ROW_ALT, fill_type="solid") \
               if row_i % 2 == 0 else None

        for col_i, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.border = _thin()
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill
            # Покупки > 0 → зелёный, 0 → обычный
            if col_i == 7 and val > 0:
                cell.font = Font(color=_GREEN, bold=True, name="Calibri")
            elif col_i == 8 and val > 0:
                cell.font = Font(color=_GREEN, bold=True, name="Calibri")

    # ── Лист «Статистика» ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Статистика")
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 18

    stat_hdr = Font(bold=True, size=12, color=_HEADER_COLOR, name="Calibri")
    total_purch = sum(p[0] for p in purch.values())
    total_stars = sum(p[1] for p in purch.values())
    buyers      = sum(1 for p in purch.values() if p[0] > 0)

    stats = [
        ("Файл обновлён",         datetime.now().strftime("%d.%m.%Y %H:%M")),
        ("", ""),
        ("Всего клиентов",        len(users)),
        ("Клиентов с покупками",  buyers),
        ("Клиентов без покупок",  len(users) - buyers),
        ("Всего покупок",         total_purch),
        ("Итого Stars заработано",total_stars),
    ]

    for r, (label, val) in enumerate(stats, 1):
        cell_a = ws2.cell(row=r, column=1, value=label)
        cell_b = ws2.cell(row=r, column=2, value=val)
        if label:
            cell_a.font = stat_hdr
            cell_b.font = Font(size=12, name="Calibri")

    # Сохраняем во временный файл рядом, потом атомарно заменяем основной
    tmp_fd, tmp_path = tempfile.mkstemp(dir=os.path.dirname(EXCEL_PATH), suffix=".xlsx")
    os.close(tmp_fd)
    try:
        wb.save(tmp_path)
        shutil.move(tmp_path, EXCEL_PATH)
    except PermissionError:
        # Основной файл открыт в Excel — сохраняем резервную копию
        backup = EXCEL_PATH.replace(".xlsx", "_обновление.xlsx")
        try:
            shutil.move(tmp_path, backup)
            log.warning("Клиенты.xlsx заблокирован Excel — данные сохранены в %s", backup)
        except Exception:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
    except Exception:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
        raise
    return len(users)


async def update_clients_excel():
    """Асинхронная обёртка — вызывай после изменений в БД."""
    try:
        loop = asyncio.get_event_loop()
        count = await loop.run_in_executor(None, _sync_excel)
        return count
    except Exception as e:
        log.error("Ошибка обновления Клиенты.xlsx: %s", e)
